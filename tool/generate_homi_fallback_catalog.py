#!/usr/bin/env python3
"""Generate HOMI fallback catalogs from the approved XLSX workbook.

The JSON catalog preserves every source row and its original fields.  The Dart
catalog is a dependency-free, synchronous lookup table for voice resolvers.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path(
    r"C:\Users\Windows\Downloads"
    r"\HOMI_Danh_muc_cau_thoai_AI_Fallback_Kho_500_cau_Gui_IT.xlsx"
)
DEFAULT_JSON_OUTPUT = (
    REPO_ROOT / "assets" / "data" / "homi_ai_fallback_catalog_v1.json"
)
DEFAULT_DART_OUTPUT = (
    REPO_ROOT
    / "lib"
    / "features"
    / "voice_navigation"
    / "domain"
    / "homi_fallback_catalog.dart"
)

SCHEMA_VERSION = "1.0.0"
SOURCE_VERSION = "HOMI_AI_FALLBACK_KHO_500_CAU_V1"
APPROVED_VALUE = "✅ THÔNG QUA"
HEADER_ROW = 4
DATA_START_ROW = HEADER_ROW + 1


class CatalogValidationError(ValueError):
    """Raised when source content drifts from the approved catalog."""


class SheetSpec:
    def __init__(
        self,
        *,
        collection: str,
        sheet_name: str,
        headers: tuple[str, ...],
        expected_count: int,
        id_header: str,
    ) -> None:
        self.collection = collection
        self.sheet_name = sheet_name
        self.headers = headers
        self.expected_count = expected_count
        self.id_header = id_header


SHEET_SPECS = (
    SheetSpec(
        collection="assistantPrompts",
        sheet_name="Câu trợ lý",
        headers=(
            "ID",
            "Nhóm luồng",
            "Trạng thái",
            "Kích hoạt",
            "Câu trợ lý HOMI (VI)",
            "Loại",
            "Kênh",
            "Hành động sau câu",
            "Mở mic lại?",
            "Nền tảng",
            "Nguồn code",
            "Dòng",
            "Ghi chú",
            "Câu gốc trong code",
            "Việc IT cần làm",
        ),
        expected_count=69,
        id_header="ID",
    ),
    SheetSpec(
        collection="childUtterances",
        sheet_name="Ý định & câu trẻ",
        headers=(
            "STT",
            "Intent ID",
            "Nhóm",
            "Trạng thái",
            "Câu chuẩn",
            "Loại biến thể",
            "Câu / cụm từ trẻ có thể nói",
            "Kết quả / điều hướng",
            "Ưu tiên",
            "Nguồn code",
            "Dòng code",
            "DUYỆT KẾT QUẢ",
            "Vùng miền & độ tuổi gợi ý",
            "Ghi chú ngôn ngữ",
            "Việc IT cần làm",
        ),
        expected_count=500,
        id_header="STT",
    ),
    SheetSpec(
        collection="silenceScenarios",
        sheet_name="Im lặng & không rõ",
        headers=(
            "ID",
            "Luồng",
            "Tình huống",
            "Mốc thời gian",
            "Câu HOMI nói",
            "Hành động",
            "Mở mic lại?",
            "Lần",
            "Nguồn code",
            "Dòng",
            "Trạng thái/ghi chú",
            "Câu gốc trong code",
            "Việc IT cần làm",
        ),
        expected_count=9,
        id_header="ID",
    ),
    SheetSpec(
        collection="fallbackPolicies",
        sheet_name="Fallback ngoài phạm vi",
        headers=(
            "ID",
            "Ngữ cảnh",
            "Điều kiện ngoài phạm vi",
            "Hành vi / lời thoại cần áp dụng",
            "Ví dụ câu trẻ",
            "Đề xuất lần 1",
            "Hành động lần 1",
            "Đề xuất lần 2",
            "Hành động cuối",
            "Ưu tiên",
            "Trạng thái",
            "Việc IT cần làm",
        ),
        expected_count=8,
        id_header="ID",
    ),
)

_PLACEHOLDER_RE = re.compile(
    r"(?P<bracket>\[(?P<bracket_name>[^\[\]]+)\])"
    r"|(?P<curly>\{(?P<curly_name>[^{}]+)\})"
)
_VIETNAMESE_REPLACEMENTS = {
    "a": "àáạảãâầấậẩẫăằắặẳẵ",
    "e": "èéẹẻẽêềếệểễ",
    "i": "ìíịỉĩ",
    "o": "òóọỏõôồốộổỗơờớợởỡ",
    "u": "ùúụủũưừứựửữ",
    "y": "ỳýỵỷỹ",
    "d": "đ",
}


def _as_json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _is_populated(values: Iterable[Any]) -> bool:
    return any(_as_text(value) for value in values)


def _normalize_placeholder_name(name: str) -> str:
    value = name.strip().lower()
    for replacement, accented in _VIETNAMESE_REPLACEMENTS.items():
        value = re.sub(f"[{accented}]", replacement, value)
    return re.sub(r"\s+", " ", value)


def _placeholder_semantic(name: str) -> str:
    normalized = _normalize_placeholder_name(name)
    if normalized == "so" or normalized.startswith("so "):
        return "number"
    if normalized in {"number", "numeric"}:
        return "number"
    return "value"


def describe_text(value: Any) -> dict[str, Any]:
    """Preserve raw text while making placeholders explicit for a resolver."""

    raw = _as_text(value)
    matches = list(_PLACEHOLDER_RE.finditer(raw))
    if not matches:
        return {"kind": "literal", "raw": raw}

    literal_segments: list[str] = []
    placeholders: list[dict[str, str]] = []
    cursor = 0
    for match in matches:
        literal_segments.append(raw[cursor : match.start()])
        notation = "bracket" if match.group("bracket") else "curly"
        name = (
            match.group("bracket_name")
            if notation == "bracket"
            else match.group("curly_name")
        )
        assert name is not None
        placeholders.append(
            {
                "raw": match.group(0),
                "name": name.strip(),
                "notation": notation,
                "semantic": _placeholder_semantic(name),
            }
        )
        cursor = match.end()
    literal_segments.append(raw[cursor:])
    has_number = any(item["semantic"] == "number" for item in placeholders)
    return {
        "kind": "numeric-template" if has_number else "template",
        "raw": raw,
        "literalSegments": literal_segments,
        "placeholders": placeholders,
    }


def _read_structured_sheet(workbook: Any, spec: SheetSpec) -> list[dict[str, Any]]:
    if spec.sheet_name not in workbook.sheetnames:
        raise CatalogValidationError(f"Missing source sheet: {spec.sheet_name}")
    sheet = workbook[spec.sheet_name]
    header_cells = next(
        sheet.iter_rows(
            min_row=HEADER_ROW,
            max_row=HEADER_ROW,
            min_col=1,
            max_col=len(spec.headers),
            values_only=True,
        )
    )
    headers = tuple(_as_text(value) for value in header_cells)
    if headers != spec.headers:
        raise CatalogValidationError(
            f"Header drift in {spec.sheet_name}: expected {spec.headers!r}, "
            f"received {headers!r}"
        )

    records: list[dict[str, Any]] = []
    for row_number, cells in enumerate(
        sheet.iter_rows(
            min_row=DATA_START_ROW,
            max_col=len(spec.headers),
            values_only=True,
        ),
        start=DATA_START_ROW,
    ):
        if not _is_populated(cells):
            continue
        fields = {
            header: _as_json_value(cells[index])
            for index, header in enumerate(spec.headers)
        }
        record_id = _as_text(fields[spec.id_header])
        if not record_id:
            raise CatalogValidationError(
                f"{spec.sheet_name} row {row_number} is missing {spec.id_header}"
            )
        records.append(
            {
                "source": {
                    "sheet": spec.sheet_name,
                    "row": row_number,
                    "id": record_id,
                },
                "sourceFields": fields,
            }
        )
    if len(records) != spec.expected_count:
        raise CatalogValidationError(
            f"Count drift in {spec.sheet_name}: expected {spec.expected_count}, "
            f"received {len(records)}"
        )
    return records


def _validate_unique_ids(records: list[dict[str, Any]], collection: str) -> None:
    seen: set[str] = set()
    duplicates: list[str] = []
    for record in records:
        record_id = record["source"]["id"]
        if record_id in seen:
            duplicates.append(record_id)
        seen.add(record_id)
    if duplicates:
        raise CatalogValidationError(
            f"Duplicate source IDs in {collection}: {', '.join(sorted(set(duplicates)))}"
        )


def validate_records(records_by_collection: dict[str, list[dict[str, Any]]]) -> None:
    """Fail fast on count, approval, source-ID, and child-row drift."""

    for collection in ("assistantPrompts", "silenceScenarios", "fallbackPolicies"):
        _validate_unique_ids(records_by_collection[collection], collection)

    children = records_by_collection["childUtterances"]
    child_row_numbers = [item["sourceFields"]["STT"] for item in children]
    if sorted(child_row_numbers) != list(range(1, len(children) + 1)):
        raise CatalogValidationError(
            "Child utterance STT values must contain every value from 1 through 500"
        )

    non_approved_rows = [
        item["source"]["row"]
        for item in children
        if _as_text(item["sourceFields"]["DUYỆT KẾT QUẢ"]) != APPROVED_VALUE
    ]
    if non_approved_rows:
        raise CatalogValidationError(
            "Approval drift in Ý định & câu trẻ; expected "
            f"{APPROVED_VALUE!r} at rows {', '.join(map(str, non_approved_rows))}"
        )

    intent_ids = OrderedDict(
        (_as_text(item["sourceFields"]["Intent ID"]), None) for item in children
    )
    if len(intent_ids) != 21 or any(not intent_id for intent_id in intent_ids):
        raise CatalogValidationError(
            f"Intent count drift: expected 21 valid IDs, received {len(intent_ids)}"
        )


def _build_assistant_prompt(record: dict[str, Any]) -> dict[str, Any]:
    fields = record["sourceFields"]
    return {
        "id": record["source"]["id"],
        "flow": fields["Nhóm luồng"],
        "state": fields["Trạng thái"],
        "trigger": fields["Kích hoạt"],
        "message": describe_text(fields["Câu trợ lý HOMI (VI)"]),
        "kind": fields["Loại"],
        "channel": fields["Kênh"],
        "postAction": fields["Hành động sau câu"],
        "reopenMic": fields["Mở mic lại?"],
        "platform": fields["Nền tảng"],
        "source": record["source"],
        "sourceFields": fields,
    }


def _build_child_utterance(record: dict[str, Any]) -> dict[str, Any]:
    fields = record["sourceFields"]
    return {
        "rowNumber": fields["STT"],
        "intentId": fields["Intent ID"],
        "group": fields["Nhóm"],
        "state": fields["Trạng thái"],
        "canonical": fields["Câu chuẩn"],
        "variantStatus": fields["Loại biến thể"],
        "utterance": describe_text(fields["Câu / cụm từ trẻ có thể nói"]),
        "outcome": fields["Kết quả / điều hướng"],
        "priority": fields["Ưu tiên"],
        "approval": fields["DUYỆT KẾT QUẢ"],
        "source": record["source"],
        "sourceFields": fields,
    }


def _build_silence_scenario(record: dict[str, Any]) -> dict[str, Any]:
    fields = record["sourceFields"]
    return {
        "id": record["source"]["id"],
        "flow": fields["Luồng"],
        "situation": fields["Tình huống"],
        "timing": fields["Mốc thời gian"],
        "message": describe_text(fields["Câu HOMI nói"]),
        "action": fields["Hành động"],
        "reopenMic": fields["Mở mic lại?"],
        "attempt": fields["Lần"],
        "source": record["source"],
        "sourceFields": fields,
    }


def _build_fallback_policy(record: dict[str, Any]) -> dict[str, Any]:
    fields = record["sourceFields"]
    return {
        "id": record["source"]["id"],
        "context": fields["Ngữ cảnh"],
        "outOfScopeCondition": fields["Điều kiện ngoài phạm vi"],
        "behavior": fields["Hành vi / lời thoại cần áp dụng"],
        "exampleChildUtterance": describe_text(fields["Ví dụ câu trẻ"]),
        "firstPrompt": describe_text(fields["Đề xuất lần 1"]),
        "firstAction": fields["Hành động lần 1"],
        "secondPrompt": describe_text(fields["Đề xuất lần 2"]),
        "finalAction": fields["Hành động cuối"],
        "priority": fields["Ưu tiên"],
        "status": fields["Trạng thái"],
        "source": record["source"],
        "sourceFields": fields,
    }


def _build_intents(children: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for child in children:
        grouped.setdefault(_as_text(child["intentId"]), []).append(child)

    intents: list[dict[str, Any]] = []
    for intent_id, variants in grouped.items():
        first = variants[0]
        for field in ("group", "state", "canonical", "outcome", "priority"):
            values = {json.dumps(item[field], ensure_ascii=False) for item in variants}
            if len(values) != 1:
                raise CatalogValidationError(
                    f"Intent {intent_id} has inconsistent {field} values"
                )
        intents.append(
            {
                "id": intent_id,
                "group": first["group"],
                "state": first["state"],
                "canonical": first["canonical"],
                "outcome": first["outcome"],
                "priority": first["priority"],
                "childRowNumbers": [item["rowNumber"] for item in variants],
                "sourceRows": [item["source"]["row"] for item in variants],
                "literalVariantCount": sum(
                    item["utterance"]["kind"] == "literal" for item in variants
                ),
                "numericTemplateCount": sum(
                    item["utterance"]["kind"] == "numeric-template"
                    for item in variants
                ),
            }
        )
    return intents


def _read_overview(workbook: Any) -> dict[str, Any]:
    sheet_name = "Tổng quan"
    if sheet_name not in workbook.sheetnames:
        raise CatalogValidationError(f"Missing source sheet: {sheet_name}")
    rows: list[dict[str, Any]] = []
    for row_number, cells in enumerate(
        workbook[sheet_name].iter_rows(values_only=True),
        start=1,
    ):
        if _is_populated(cells):
            rows.append(
                {
                    "row": row_number,
                    "cells": [_as_json_value(value) for value in cells],
                }
            )
    return {"sheet": sheet_name, "rows": rows}


def build_catalog(workbook_path: Path) -> dict[str, Any]:
    if not workbook_path.is_file():
        raise CatalogValidationError(f"Source workbook does not exist: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        records = {
            spec.collection: _read_structured_sheet(workbook, spec)
            for spec in SHEET_SPECS
        }
        validate_records(records)
        assistant_prompts = [
            _build_assistant_prompt(item) for item in records["assistantPrompts"]
        ]
        child_utterances = [
            _build_child_utterance(item) for item in records["childUtterances"]
        ]
        silence_scenarios = [
            _build_silence_scenario(item) for item in records["silenceScenarios"]
        ]
        fallback_policies = [
            _build_fallback_policy(item) for item in records["fallbackPolicies"]
        ]
        intents = _build_intents(child_utterances)
        literal_count = sum(
            item["utterance"]["kind"] == "literal" for item in child_utterances
        )
        numeric_count = sum(
            item["utterance"]["kind"] == "numeric-template"
            for item in child_utterances
        )
        return {
            "schemaVersion": SCHEMA_VERSION,
            "sourceVersion": SOURCE_VERSION,
            "source": {
                "workbook": workbook_path.name,
                "headerRow": HEADER_ROW,
                "sheets": {
                    spec.sheet_name: {
                        "collection": spec.collection,
                        "headers": list(spec.headers),
                        "expectedRecordCount": spec.expected_count,
                    }
                    for spec in SHEET_SPECS
                },
            },
            "overview": _read_overview(workbook),
            "assistantPrompts": assistant_prompts,
            "childUtterances": child_utterances,
            "intents": intents,
            "silenceScenarios": silence_scenarios,
            "fallbackPolicies": fallback_policies,
            "validation": {
                "approvedChildUtteranceValue": APPROVED_VALUE,
                "approvedChildUtteranceCount": len(child_utterances),
                "assistantPromptCount": len(assistant_prompts),
                "childUtteranceCount": len(child_utterances),
                "intentCount": len(intents),
                "literalChildPhraseCount": literal_count,
                "numericChildTemplateCount": numeric_count,
                "silenceScenarioCount": len(silence_scenarios),
                "fallbackPolicyCount": len(fallback_policies),
                "sourceNotices": [
                    "Source overview claims 18 silence/fallback scenarios; "
                    "the source tabs contain 9 + 8 = 17 records.",
                    "Non-contiguous source IDs are intentionally preserved: "
                    "AI-021, INT-005, FB-002, and FB-003 are absent.",
                ],
            },
        }
    finally:
        workbook.close()


def _dart_string(value: Any) -> str:
    text = _as_text(value)
    escaped = (
        text.replace("\\", "\\\\")
        .replace("'", "\\'")
        .replace("$", "\\$")
        .replace("\r", "\\r")
        .replace("\n", "\\n")
    )
    return f"'{escaped}'"


def _literal_and_numeric_variants(
    children: list[dict[str, Any]],
) -> tuple[OrderedDict[str, list[str]], OrderedDict[str, list[dict[str, Any]]]]:
    literals: OrderedDict[str, list[str]] = OrderedDict()
    numeric_templates: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for child in children:
        intent_id = _as_text(child["intentId"])
        phrase = child["utterance"]
        if phrase["kind"] == "literal":
            literals.setdefault(intent_id, []).append(phrase["raw"])
        elif phrase["kind"] == "numeric-template":
            numeric_templates.setdefault(intent_id, []).append(phrase)
    return literals, numeric_templates


def build_dart_catalog(catalog: dict[str, Any]) -> str:
    literals, numeric_templates = _literal_and_numeric_variants(
        catalog["childUtterances"]
    )
    lines = [
        "// GENERATED CODE - DO NOT MODIFY BY HAND.",
        "// Source: HOMI_Danh_muc_cau_thoai_AI_Fallback_Kho_500_cau_Gui_IT.xlsx",
        "// Regenerate with: python tool/generate_homi_fallback_catalog.py",
        "",
        "class HomiNumericPhrasePattern {",
        "  const HomiNumericPhrasePattern({",
        "    required this.raw,",
        "    required this.literalSegments,",
        "    required this.placeholderNames,",
        "  });",
        "",
        "  final String raw;",
        "  final List<String> literalSegments;",
        "  final List<String> placeholderNames;",
        "}",
        "",
        "class HomiFallbackPolicy {",
        "  const HomiFallbackPolicy({",
        "    required this.id,",
        "    required this.context,",
        "    required this.condition,",
        "    required this.behavior,",
        "    required this.exampleChildUtterance,",
        "    required this.firstPrompt,",
        "    required this.firstAction,",
        "    required this.secondPrompt,",
        "    required this.finalAction,",
        "    required this.priority,",
        "    required this.status,",
        "  });",
        "",
        "  final String id;",
        "  final String context;",
        "  final String condition;",
        "  final String behavior;",
        "  final String exampleChildUtterance;",
        "  final String firstPrompt;",
        "  final String firstAction;",
        "  final String secondPrompt;",
        "  final String finalAction;",
        "  final String priority;",
        "  final String status;",
        "}",
        "",
        "/// Dependency-free, synchronous copy of the approved HOMI corpus.",
        "class HomiFallbackCatalog {",
        "  const HomiFallbackCatalog._();",
        "",
        "  static const int literalChildPhraseCount = "
        f"{catalog['validation']['literalChildPhraseCount']};",
        "  static const int numericChildTemplateCount = "
        f"{catalog['validation']['numericChildTemplateCount']};",
        "",
        "  /// Literal phrases only. Numeric bracket templates remain separate so",
        "  /// callers activate them only while a number is expected.",
        "  static const Map<String, List<String>> childPhrasesByIntent =",
        "      <String, List<String>>{",
    ]
    for intent_id, phrases in literals.items():
        lines.append(f"        {_dart_string(intent_id)}: <String>[")
        lines.extend(f"          {_dart_string(phrase)}," for phrase in phrases)
        lines.append("        ],")
    lines.extend(
        [
            "      };",
            "",
            "  static const Map<String, List<HomiNumericPhrasePattern>>",
            "  numericChildPatternsByIntent =",
            "      <String, List<HomiNumericPhrasePattern>>{",
        ]
    )
    for intent_id, patterns in numeric_templates.items():
        lines.append(
            f"        {_dart_string(intent_id)}: <HomiNumericPhrasePattern>["
        )
        for pattern in patterns:
            lines.extend(
                [
                    "          HomiNumericPhrasePattern(",
                    f"            raw: {_dart_string(pattern['raw'])},",
                    "            literalSegments: <String>[",
                ]
            )
            lines.extend(
                f"              {_dart_string(segment)},"
                for segment in pattern["literalSegments"]
            )
            lines.extend(["            ],", "            placeholderNames: <String>["])
            lines.extend(
                f"              {_dart_string(item['name'])},"
                for item in pattern["placeholders"]
            )
            lines.extend(["            ],", "          ),"])
        lines.append("        ],")
    lines.extend(
        [
            "      };",
            "",
            "  static const Map<String, String> assistantPromptById =",
            "      <String, String>{",
        ]
    )
    for prompt in catalog["assistantPrompts"]:
        lines.append(
            f"        {_dart_string(prompt['id'])}: "
            f"{_dart_string(prompt['message']['raw'])},"
        )
    lines.extend(
        [
            "      };",
            "",
            "  static const Map<String, String> silencePromptById =",
            "      <String, String>{",
        ]
    )
    for scenario in catalog["silenceScenarios"]:
        lines.append(
            f"        {_dart_string(scenario['id'])}: "
            f"{_dart_string(scenario['message']['raw'])},"
        )
    lines.extend(
        [
            "      };",
            "",
            "  static const Map<String, HomiFallbackPolicy> fallbackPolicyById =",
            "      <String, HomiFallbackPolicy>{",
        ]
    )
    for policy in catalog["fallbackPolicies"]:
        lines.extend(
            [
                f"        {_dart_string(policy['id'])}: HomiFallbackPolicy(",
                f"          id: {_dart_string(policy['id'])},",
                f"          context: {_dart_string(policy['context'])},",
                f"          condition: {_dart_string(policy['outOfScopeCondition'])},",
                f"          behavior: {_dart_string(policy['behavior'])},",
                "          exampleChildUtterance: "
                f"{_dart_string(policy['exampleChildUtterance']['raw'])},",
                f"          firstPrompt: {_dart_string(policy['firstPrompt']['raw'])},",
                f"          firstAction: {_dart_string(policy['firstAction'])},",
                f"          secondPrompt: {_dart_string(policy['secondPrompt']['raw'])},",
                f"          finalAction: {_dart_string(policy['finalAction'])},",
                f"          priority: {_dart_string(policy['priority'])},",
                f"          status: {_dart_string(policy['status'])},",
                "        ),",
            ]
        )
    lines.extend(
        [
            "      };",
            "",
            "  /// Case-, accent-, and punctuation-insensitive normalization.",
            "  static String normalizeVietnamese(String value) {",
            "    var normalized = value.trim().toLowerCase();",
            "    const replacements = <String, String>{",
            "      'a': 'àáạảãâầấậẩẫăằắặẳẵ',",
            "      'e': 'èéẹẻẽêềếệểễ',",
            "      'i': 'ìíịỉĩ',",
            "      'o': 'òóọỏõôồốộổỗơờớợởỡ',",
            "      'u': 'ùúụủũưừứựửữ',",
            "      'y': 'ỳýỵỷỹ',",
            "      'd': 'đ',",
            "    };",
            "    for (final entry in replacements.entries) {",
            "      normalized = normalized.replaceAll(",
            "        RegExp('[" + "$" + "{entry.value}]'),",
            "        entry.key,",
            "      );",
            "    }",
            "    return normalized",
            "        .replaceAll(RegExp(r'[^a-z0-9]+'), ' ')",
            "        .replaceAll(RegExp(r'\\s+'), ' ')",
            "        .trim();",
            "  }",
            "",
            "  /// Returns true only when the entire normalized transcript matches.",
            "  static bool matchesWholePhrase(String transcript, String phrase) {",
            "    final normalizedTranscript = normalizeVietnamese(transcript);",
            "    final normalizedPhrase = normalizeVietnamese(phrase);",
            "    return normalizedTranscript.isNotEmpty &&",
            "        normalizedTranscript == normalizedPhrase;",
            "  }",
            "",
            "  static bool matchesChildPhrase(String intentId, String transcript) {",
            "    final phrases = childPhrasesByIntent[intentId] ?? const <String>[];",
            "    return phrases.any(",
            "      (phrase) => matchesWholePhrase(transcript, phrase),",
            "    );",
            "  }",
            "}",
            "",
        ]
    )
    return "\n".join(lines)


def write_catalogs(
    catalog: dict[str, Any], json_output: Path, dart_output: Path
) -> None:
    json_output.parent.mkdir(parents=True, exist_ok=True)
    dart_output.parent.mkdir(parents=True, exist_ok=True)
    json_output.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    dart_output.write_text(build_dart_catalog(catalog), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate HOMI fallback JSON and static Dart catalogs."
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--json-output", type=Path, default=DEFAULT_JSON_OUTPUT)
    parser.add_argument("--dart-output", type=Path, default=DEFAULT_DART_OUTPUT)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        catalog = build_catalog(args.source)
        write_catalogs(catalog, args.json_output, args.dart_output)
    except CatalogValidationError as error:
        print(f"Catalog generation failed: {error}", file=sys.stderr)
        return 1
    validation = catalog["validation"]
    print(
        "Generated HOMI fallback catalog: "
        f"{validation['assistantPromptCount']} assistant prompts, "
        f"{validation['childUtteranceCount']} child utterances / "
        f"{validation['intentCount']} intents, "
        f"{validation['silenceScenarioCount']} silence scenarios, "
        f"{validation['fallbackPolicyCount']} fallback policies; "
        f"{validation['literalChildPhraseCount']} literal phrases and "
        f"{validation['numericChildTemplateCount']} numeric templates."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
