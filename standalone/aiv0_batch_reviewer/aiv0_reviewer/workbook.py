from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import AuditEntry, AudioRecord, STANDARD_COLUMNS


HEADER_ALIASES = {
    "id": "record_id",
    "record id": "record_id",
    "mã bản ghi": "record_id",
    "ma ban ghi": "record_id",
    "batch": "batch_id",
    "batch id": "batch_id",
    "mã đợt": "batch_id",
    "ma dot": "batch_id",
    "sentence code": "sentence_code",
    "mã câu": "sentence_code",
    "ma cau": "sentence_code",
    "sentence": "source_vi",
    "text": "source_vi",
    "câu tiếng việt": "source_vi",
    "cau tieng viet": "source_vi",
    "english": "expected_en",
    "expected english": "expected_en",
    "tiếng anh mong đợi": "expected_en",
    "draft file": "draft_filename",
    "filename": "draft_filename",
    "tên file": "draft_filename",
    "official file": "official_filename",
    "audio id": "audio_id",
    "status": "upload_status",
    "trạng thái": "upload_status",
    "kết quả kiểm thử": "test_result",
    "ket qua kiem thu": "test_result",
    "loại lỗi": "error_categories",
    "loai loi": "error_categories",
    "kết quả thực tế": "observed_result",
    "đề xuất sửa": "repair_suggestion",
    "de xuat sua": "repair_suggestion",
    "trạng thái sửa": "repair_status",
    "ghi chú": "note",
}


def _canonical_header(value: Any) -> str:
    header = "" if value is None else str(value).strip()
    lowered = header.casefold().replace("_", " ").replace("-", " ")
    lowered = " ".join(lowered.split())
    if lowered in HEADER_ALIASES:
        return HEADER_ALIASES[lowered]
    snake = lowered.replace(" ", "_")
    return snake or header


def _validate(records: list[AudioRecord]) -> None:
    if not records:
        raise ValueError("Bảng không có bản ghi dữ liệu.")
    missing = [index + 2 for index, record in enumerate(records) if not record.record_id]
    if missing:
        raise ValueError(f"Thiếu record_id tại dòng: {', '.join(map(str, missing[:10]))}")
    seen: set[str] = set()
    duplicates: list[str] = []
    for record in records:
        key = record.record_id.casefold()
        if key in seen:
            duplicates.append(record.record_id)
        seen.add(key)
    if duplicates:
        raise ValueError(f"record_id bị trùng: {', '.join(duplicates[:10])}")


def load_records(path: str | Path) -> list[AudioRecord]:
    source = Path(path)
    suffix = source.suffix.casefold()
    if suffix == ".xlsx":
        records = _load_xlsx(source)
    elif suffix in {".csv", ".tsv"}:
        records = _load_delimited(source, "\t" if suffix == ".tsv" else ",")
    else:
        raise ValueError("Chỉ hỗ trợ .xlsx, .csv và .tsv.")
    _validate(records)
    return records


def _load_xlsx(path: Path) -> list[AudioRecord]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = workbook["Samples"] if "Samples" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        return []
    headers = [_canonical_header(value) for value in raw_headers]
    result: list[AudioRecord] = []
    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue
        values = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        result.append(AudioRecord.from_mapping(values))
    workbook.close()
    return result


def _load_delimited(path: Path, delimiter: str) -> list[AudioRecord]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        result = []
        for row in reader:
            values = {_canonical_header(key): value for key, value in row.items() if key}
            if any(value not in (None, "") for value in values.values()):
                result.append(AudioRecord.from_mapping(values))
        return result


def save_workbook(
    path: str | Path,
    records: Iterable[AudioRecord],
    audit_entries: Iterable[AuditEntry] = (),
) -> None:
    record_list = list(records)
    extras = sorted({key for record in record_list for key in record.extra})
    headers = list(STANDARD_COLUMNS) + extras

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Samples"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(record_list) + 1)}"
    sheet.append(headers)
    for record in record_list:
        values = record.to_mapping()
        sheet.append([values.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="173F5F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")

    widths = {
        "record_id": 19,
        "batch_id": 14,
        "sentence_code": 14,
        "source_vi": 42,
        "expected_en": 42,
        "draft_filename": 28,
        "official_filename": 28,
        "local_audio_path": 46,
        "audio_id": 24,
        "audio_hash": 68,
        "audio_version": 14,
        "upload_status": 20,
        "uploaded_at": 26,
        "test_result": 18,
        "error_categories": 38,
        "observed_result": 42,
        "repair_suggestion": 46,
        "repair_status": 18,
        "reviewer": 18,
        "reviewed_at": 26,
        "note": 36,
    }
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(header, 18)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    audit_sheet = workbook.create_sheet("Upload Log")
    audit_headers = ["timestamp", "record_id", "action", "old_value", "new_value", "detail"]
    audit_sheet.append(audit_headers)
    for entry in audit_entries:
        audit_sheet.append(
            [entry.timestamp, entry.record_id, entry.action, entry.old_value, entry.new_value, entry.detail]
        )
    for cell in audit_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="206A5D")
        cell.font = Font(color="FFFFFF", bold=True)
    audit_sheet.freeze_panes = "A2"
    for index, width in enumerate((26, 20, 24, 34, 34, 52), start=1):
        audit_sheet.column_dimensions[get_column_letter(index)].width = width

    guide = workbook.create_sheet("Instructions")
    guide.append(["AIV0 Batch Audio Manager"])
    guide.append(["1. record_id là mã cố định, không thay đổi khi đổi audio."])
    guide.append(["2. draft_filename dùng để ghép file audio trong thư mục."])
    guide.append(["3. audio_id được tạo cục bộ hoặc nhận từ API upload chính thức."])
    guide.append(["4. Không ghi tên thật hoặc thông tin riêng của trẻ vào bảng."])
    guide.append(["5. Mẫu Không đạt phải có ít nhất một loại lỗi trước khi lưu."])
    guide.append(["6. Sửa hàng loạt chỉ áp dụng cho nhóm Không đạt có cùng loại lỗi."])
    guide.column_dimensions["A"].width = 100
    guide["A1"].font = Font(size=16, bold=True, color="173F5F")

    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def create_template(path: str | Path, count: int = 207) -> None:
    records = [
        AudioRecord(
            record_id=f"CV26-TEEN-{index:03d}",
            batch_id="CV26-TEEN",
            sentence_code=f"S{index:03d}",
        )
        for index in range(1, count + 1)
    ]
    save_workbook(path, records)
